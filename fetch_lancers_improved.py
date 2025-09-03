#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import asyncio
import aiohttp
import json
import re
import os
import openpyxl
from pathlib import Path
from typing import List
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from playwright.async_api import async_playwright

# =============================
# 環境判定
# =============================
IS_GITHUB_ACTIONS = os.getenv('GITHUB_ACTIONS') == 'true'

# =============================
# 設定値
# =============================
LANCERS_SEARCH_URL = "https://www.lancers.jp/work/search/system?budget_from=&budget_to=&work_rank%5B%5D=&work_rank%5B%5D=&work_rank%5B%5D=&keyword=&sort=work_post_date"
MAX_JOBS_TO_FETCH = 100
HEADLESS_MODE = os.getenv("HEADLESS", "true").lower() != "false"  # 環境変数で上書き可

# =============================
# 保存先（環境により切替）＋上書き対応
# =============================
def _ensure_parent_dir(p: str):
    try:
        Path(p).parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

def get_excel_path_local() -> str:
    paths = [
        r"C:\Users\SUZUKI Natsumi\株式会社ReySolid\【B】IT-Solution@ReySolid - 案件管理 - ドキュメント\案件管理\案件情報.xlsx",
        r"C:\Users\SUZUKI Natsumi\OneDrive - 株式会社ReySolid\案件情報.xlsx",
    ]
    for path in paths:
        if os.path.exists(path):
            print(f"📊 使用するExcelファイル: {path}")
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(path))
                print(f"   最終更新: {mtime}")
            except Exception:
                pass
            return path
    # 見つからない場合も第1候補を返し、後続で自動作成
    _ensure_parent_dir(paths[0])
    return paths[0]

# 優先度: 環境変数 EXCEL_PATH > 自動判定
env_override = os.getenv("EXCEL_PATH")

if IS_GITHUB_ACTIONS:
    EXCEL_PATH = env_override or "案件情報.xlsx"  # Actions はリポジトリ直下
    print("📍 GitHub Actions環境で実行中: Excelはリポジトリに保存します")
else:
    EXCEL_PATH = env_override or get_excel_path_local()
    print("📍 ローカル環境で実行中: OneDrive/SharePoint パスを使用します")

print(f"📄 EXCEL_PATH = {EXCEL_PATH}")

# =============================
# スキル設定 / 除外キーワード
# =============================
COMPANY_SKILLS = {
    "超高優先度": ["AI", "GPT", "ChatGPT", "Python", "API", "Django", "Next.js","React","TypeScript", "機械学習"],
    "高優先度": ["bot", "Talend", "Java", "スマホアプリ", "モバイル開発", "人工知能"],
    "中優先度": ["効率化", "ツール", "開発", "システム開発", "React", "Node.js", "自動化", "スクレイピング"],
    "低優先度": ["PostgreSQL", "MySQL", "社内ツール", "業務改善", "アプリ", "サイト", "管理"],
    "最低優先度": ["Render", "ロリッポップ", "WordPress", "PHP"]
}

EXCLUDE_KEYWORDS = [
    "求人", "採用", "転職", "正社員", "アルバイト", "派遣",
    "コンペ", "コンペティション", "コンテスト",
    "募集終了", "終了", "締切", "CAD",
    'デザイン', 'イラスト', 'ロゴ', '動画編集', '写真', '撮影', 'Photoshop',
    'Illustrator', 'After Effects', 'Premiere', 'グラフィック', 'UI/UX',
    'XD', 'Sketch', 'Canva', 'バナー', 'チラシ', 'ポスター',
    '求人', '採用', 'リクルート', 'スカウト', '人材', 'コンペ', 'コンテスト',
    '懸賞', '応募者多数', '選考', '審査', 'ライティング', '記事作成', 'ゲーム',
    'ブログ記事', 'SEO記事', 'コピーライティング', 'シナリオ', '脚本',
    '翻訳', '通訳', 'アンケート', 'モニター', 'レビュー',
    'テレアポ', '営業', 'カスタマーサポート', 'サポート業務', '事務',
    '経理', '秘書', 'アシスタント', '内職', '簡単作業', '軽作業'
]

# =============================
# Excel ヘルパ
# =============================
def _format_skill_matches_compact_for_excel(skill_matches: List[dict]) -> str:
    if not skill_matches:
        return ""
    ultra = [m.get("skill") for m in skill_matches if m.get("priority") == "超高優先度"]
    high  = [m.get("skill") for m in skill_matches if m.get("priority") == "高優先度"]
    mid   = [m.get("skill") for m in skill_matches if m.get("priority") == "中優先度"]
    low   = [m.get("skill") for m in skill_matches if m.get("priority") == "低優先度"]
    lowst = [m.get("skill") for m in skill_matches if m.get("priority") == "最低優先度"]
    parts = []
    if ultra: parts.append("🔥" + ",".join(ultra[:2]))
    if high:  parts.append("★" + ",".join(high[:2]))
    if mid:   parts.append("◆" + ",".join(mid[:2]))
    if low:   parts.append("◇" + ",".join(low[:1]))
    if lowst: parts.append("○" + ",".join(lowst[:1]))
    s = " ".join(parts)
    return (s[:97] + "...") if len(s) > 100 else s

def _autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

def _ensure_book_and_sheets(path: Path):
    if not path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ランサーズ"
        ws.append(["取得日時","タイトル","カテゴリ","価格","締切","URL","優先度スコア","スキル概要"])
        stat = wb.create_sheet("統計")
        stat.append(["timestamp","count","type","skill_match_rate","no_skill_match","multi_skill_match","high_priority"])
        wb.save(path)
    else:
        wb = openpyxl.load_workbook(path)
        if "ランサーズ" not in wb.sheetnames:
            ws = wb.create_sheet("ランサーズ")
            ws.append(["取得日時","タイトル","カテゴリ","価格","締切","URL","優先度スコア","スキル概要"])
        if "統計" not in wb.sheetnames:
            stat = wb.create_sheet("統計")
            stat.append(["timestamp","count","type","skill_match_rate","no_skill_match","multi_skill_match","high_priority"])
    return wb

def replace_lancers_sheet(data: dict, excel_path: str = EXCEL_PATH):
    try:
        path = Path(excel_path)
        wb = _ensure_book_and_sheets(path)

        if "ランサーズ" in wb.sheetnames:
            del wb["ランサーズ"]

        ws = wb.create_sheet("ランサーズ", 0)
        ws.append(["取得日時","タイトル","カテゴリ","価格","締切","URL","優先度スコア","スキル概要"])

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for job in data.get("jobs", []):
            url = job.get("link", "")
            ws.append([
                now_str,
                job.get("title", ""),
                job.get("category", ""),
                job.get("price", ""),
                job.get("deadline", ""),
                url,
                job.get("priority_score", ""),
                _format_skill_matches_compact_for_excel(job.get("skill_matches", []))
            ])
            last = ws.max_row
            if url:
                ws.cell(row=last, column=6).hyperlink = url
                ws.cell(row=last, column=6).style = "Hyperlink"

        _autosize_columns(ws)
        wb.save(path)
        print(f"✅ 『ランサーズ』シートを上書き保存しました: {excel_path}")

    except Exception as e:
        print(f"❌ ランサーズ上書きエラー: {e}")


# =============================
# 取得・通知クラス
# =============================
class CompleteJobsNotifier:
    def __init__(self):
        self.jobs_data = []
        self.seen_links = set()

    async def fetch_jobs(self):
        print("🚀 Lancers全案件取得を開始...")
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=HEADLESS_MODE, slow_mo=300)
            try:
                context = await browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    locale="ja-JP"
                )
                page = await context.new_page()
                print(f"📡 アクセス中: {LANCERS_SEARCH_URL}")
                response = await page.goto(LANCERS_SEARCH_URL, wait_until="domcontentloaded", timeout=60000)
                print(f"✅ ページ読み込み完了 (ステータス: {response.status})")

                await page.wait_for_timeout(5000)
                await self.scroll_and_load_more(page)

                job_elements = await page.query_selector_all("a[href*='/work/detail/']")
                print(f"📊 {len(job_elements)} 個の案件候補を発見")

                all_jobs = []
                for element in job_elements[:MAX_JOBS_TO_FETCH]:
                    job_info = await self.extract_job_info(element, page)
                    if job_info and self.should_include_job_minimal(job_info):
                        all_jobs.append(job_info)
                        skill_info = self.format_skill_matches(job_info["skill_matches"])
                        print(f"📝 案件 {len(all_jobs)}: {job_info['title'][:40]}... | {skill_info}")

                sorted_jobs = self.sort_by_skill_relevance(all_jobs)
                print(f"✅ 全 {len(sorted_jobs)} 件の案件を取得しました")
                self.jobs_data = sorted_jobs
                return sorted_jobs

            except Exception as e:
                print(f"❌ エラー: {e}")
                return []
            finally:
                await browser.close()

    async def scroll_and_load_more(self, page):
        try:
            for _ in range(5):
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(2000)
            more_button = await page.query_selector(".more-button, .load-more, [class*='more']")
            if more_button:
                await more_button.click()
                await page.wait_for_timeout(3000)
            for _ in range(3):
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(2000)
        except Exception as e:
            print(f"⚠️ スクロール読み込みエラー: {e}")

    async def extract_job_info(self, element, page):
        try:
            title_text = await element.text_content()
            href = await element.get_attribute("href")
            if not title_text or not href:
                return None
            title = self.clean_title(title_text)
            if not title or len(title) < 5:
                return None
            if href.startswith("/"):
                href = "https://www.lancers.jp" + href
            if href in self.seen_links:
                return None
            self.seen_links.add(href)

            recruitment_info = await self.extract_recruitment_details(element)
            skill_matches = self.find_all_skill_matches(title)
            job_info = {
                "title": title,
                "link": href,
                "price": recruitment_info["price"],
                "deadline": recruitment_info["deadline"],
                "applicant_count": recruitment_info["applicant_count"],
                "recruitment_count": recruitment_info["recruitment_count"],
                "client_name": recruitment_info["client_name"],
                "status": recruitment_info["status"],
                "urgency": recruitment_info["urgency"],
                "category": recruitment_info["category"],
                "skill_matches": skill_matches,
                "skill_count": len(skill_matches),
                "priority_score": self.calculate_comprehensive_score(title, recruitment_info, skill_matches),
                "scraped_at": datetime.now().isoformat()
            }
            return job_info
        except Exception as e:
            print(f"⚠️ 案件抽出エラー: {e}")
            return None

    def find_all_skill_matches(self, title):
        matches = []
        title_lower = title.lower()
        for priority, skills in COMPANY_SKILLS.items():
            for skill in skills:
                if skill.lower() in title_lower:
                    matches.append({"skill": skill, "priority": priority})
        additional_keywords = {
            "自動化": "中優先度",
            "スクレイピング": "中優先度",
            "アプリ": "中優先度",
            "サイト": "低優先度",
            "管理": "低優先度",
            "コンサル": "中優先度",
            "Ai": "超高優先度",
            "人工知能": "高優先度"
        }
        for keyword, priority in additional_keywords.items():
            if keyword.lower() in title_lower:
                matches.append({"skill": keyword, "priority": priority})
        seen_skills = set()
        unique_matches = []
        for m in matches:
            if m["skill"] not in seen_skills:
                unique_matches.append(m)
                seen_skills.add(m["skill"])
        return unique_matches

    def format_skill_matches(self, skill_matches):
        if not skill_matches:
            return "🔧 スキルセット: なし"
        skills_by_priority = {}
        for m in skill_matches:
            p = m["priority"]
            skills_by_priority.setdefault(p, []).append(m["skill"])
        parts = []
        if "超高優先度" in skills_by_priority: parts.append(f"🔥{', '.join(skills_by_priority['超高優先度'])}")
        if "高優先度" in skills_by_priority:  parts.append(f"★{', '.join(skills_by_priority['高優先度'])}")
        if "中優先度" in skills_by_priority:  parts.append(f"◆{', '.join(skills_by_priority['中優先度'])}")
        if "低優先度" in skills_by_priority:  parts.append(f"◇{', '.join(skills_by_priority['低優先度'])}")
        if "最低優先度" in skills_by_priority: parts.append(f"○{', '.join(skills_by_priority['最低優先度'])}")
        return f"🔧 スキルセット: {' | '.join(parts)}" if parts else "🔧 スキルセット: なし"

    def format_skill_matches_compact(self, skill_matches: List[dict]) -> str:
        if not skill_matches:
            return "🔧 スキルセット: なし"
        ultra = [m.get("skill") for m in skill_matches if m.get("priority") == "超高優先度"]
        high  = [m.get("skill") for m in skill_matches if m.get("priority") == "高優先度"]
        mid   = [m.get("skill") for m in skill_matches if m.get("priority") == "中優先度"]
        low   = [m.get("skill") for m in skill_matches if m.get("priority") == "低優先度"]
        lowst = [m.get("skill") for m in skill_matches if m.get("priority") == "最低優先度"]
        parts = []
        if ultra: parts.append("🔥" + ",".join(ultra[:2]))
        if high:  parts.append("★" + ",".join(high[:2]))
        if mid:   parts.append("◆" + ",".join(mid[:2]))
        if low:   parts.append("◇" + ",".join(low[:1]))
        if lowst: parts.append("○" + ",".join(lowst[:1]))
        s = " ".join(parts) if parts else ""
        if len(s) > 100: s = s[:97] + "..."
        return f"🔧 スキルセット: {s}" if s else "🔧 スキルセット: なし"

    async def extract_recruitment_details(self, element):
        recruitment_info = {
            "price": "価格情報なし",
            "deadline": "期限情報なし",
            "applicant_count": "0",
            "recruitment_count": "1",
            "client_name": "依頼者情報なし",
            "status": "募集中",
            "urgency": False,
            "category": "システム開発"
        }
        try:
            parent = await element.evaluate_handle("el => el.closest('.c-media, .p-jobList__item, article, li')")
            if parent:
                price_elem = await parent.query_selector(".c-media__price, .price, .budget, [class*='price']")
                if price_elem:
                    price_text = await price_elem.text_content()
                    if price_text and "円" in price_text:
                        recruitment_info["price"] = self.clean_price_text(price_text)
                deadline_elem = await parent.query_selector(".c-media__deadline, .deadline, [class*='deadline']")
                if deadline_elem:
                    deadline_text = await deadline_elem.text_content()
                    if deadline_text:
                        recruitment_info["deadline"] = deadline_text.strip()
                        if any(w in deadline_text for w in ["急募", "緊急", "即日", "至急"]):
                            recruitment_info["urgency"] = True
                applicant_elem = await parent.query_selector(".c-media__applicant, .applicant, [class*='applicant']")
                if applicant_elem:
                    applicant_text = await applicant_elem.text_content()
                    if applicant_text:
                        numbers = re.findall(r'(\d+)', applicant_text)
                        if len(numbers) >= 2:
                            recruitment_info["applicant_count"] = numbers[0]
                            recruitment_info["recruitment_count"] = numbers[1]
        except:
            pass
        return recruitment_info

    def clean_title(self, title):
        if not title:
            return ""
        try:
            import unicodedata
            title = unicodedata.normalize('NFKC', title)
        except:
            pass
        title = re.sub(r'\s+', ' ', title.strip())
        title = re.sub(r'^(NEW\s*){1,}', '', title, flags=re.IGNORECASE)
        title = re.sub(r'^\d+回目\s*', '', title)
        return title.strip()

    def clean_price_text(self, raw_price):
        if not raw_price:
            return "価格情報なし"
        try:
            import unicodedata
            raw_price = unicodedata.normalize('NFKC', raw_price)
        except:
            pass
        price = re.sub(r'\s+', ' ', raw_price.strip())
        price = re.sub(r'円\s*/\s*', '円 / ', price)
        return price

    def calculate_comprehensive_score(self, title, recruitment_info, skill_matches):
        score = 0
        title_lower = title.lower()
        for m in skill_matches:
            p = m["priority"]
            if p == "超高優先度": score += 100
            elif p == "高優先度": score += 50
            elif p == "中優先度": score += 20
            elif p == "低優先度": score += 10
            elif p == "最低優先度": score += 5
        if len(skill_matches) >= 3: score += 50
        elif len(skill_matches) >= 2: score += 25
        elif len(skill_matches) >= 1: score += 10
        priority_keywords = {
            "chatgpt": 80, "python": 70, "api": 60, "ai": 60,
            "自動化": 40, "bot": 40, "効率化": 30, "ツール": 25, "開発": 20, "システム": 15
        }
        for k, bonus in priority_keywords.items():
            if k in title_lower:
                score += bonus
        if recruitment_info["urgency"]:
            score += 15
        try:
            applicant_count = int(recruitment_info["applicant_count"])
            if applicant_count == 0: score += 10
            elif applicant_count <= 2: score += 5
        except:
            pass
        price_text = recruitment_info["price"]
        if "円" in price_text:
            numbers = re.findall(r'(\d+,?\d*)', price_text.replace(',', ''))
            if numbers:
                try:
                    max_price = max([int(num.replace(',', '')) for num in numbers])
                    if max_price >= 500000: score += 15
                    elif max_price >= 100000: score += 8
                    elif max_price >= 50000:  score += 3
                except:
                    pass
        return score

    def should_include_job_minimal(self, job_info):
        title = job_info["title"]
        if not title or len(title.strip()) < 5:
            return False
        title_lower = title.lower()
        for keyword in EXCLUDE_KEYWORDS:
            if keyword.lower() in title_lower:
                return False
        status = job_info["status"]
        if any(w in status for w in ["募集終了", "締切", "終了", "完了"]):
            return False
        if job_info["priority_score"] >= 10 or job_info["skill_count"] >= 1:
            return True
        priority_keywords = ["chatgpt", "python", "api", "ai", "自動化", "bot", "効率化", "ツール", "開発", "システム"]
        if any(k in title_lower for k in priority_keywords):
            return True
        return False

    def sort_by_skill_relevance(self, jobs):
        def sort_key(job):
            base_score = job["priority_score"]
            if job["skill_count"] == 0:
                base_score -= 1000
            applicant_count = int(job["applicant_count"]) if job["applicant_count"].isdigit() else 999
            return (-base_score, -job["skill_count"], applicant_count, not job["urgency"], job["scraped_at"])
        return sorted(jobs, key=sort_key)

    def create_teams_payload(self, jobs):
        if not jobs:
            return {
                "@type": "MessageCard",
                "@context": "https://schema.org/extensions",
                "summary": "Lancers全案件通知",
                "themeColor": "0078D4",
                "title": "🚀 Lancers全案件リスト",
                "text": "📭 現在条件に合う案件が見つかりませんでした。"
            }
        MAX_CHARS = 25000
        header_text = f"現在の全案件リストです（**{len(jobs)}件**を発見）\n\n"
        footer_text = "\n\n📋 詳細情報はJSONファイルでも確認できます。"
        available_chars = MAX_CHARS - len(header_text) - len(footer_text) - 500
        main_content = ""
        displayed_count = 0
        skipped_count = 0
        print(f"📊 Teams表示制限: {MAX_CHARS:,}文字まで利用可能")
        for i, job in enumerate(jobs, 1):
            skill_info = self.format_skill_matches_compact(job["skill_matches"])
            job_text  = f"**{i}. {job['title']}**  \n"
            job_text += f"💰 {job['price']}  \n"
            if job['deadline'] != "期限情報なし" and len(job['deadline']) < 20:
                job_text += f"⏰ {job['deadline']}  \n"
            if job['applicant_count'] != "0":
                job_text += f"👥 応募{job['applicant_count']}人  \n"
            job_text += f"{skill_info}  \n"
            if job['urgency']:
                job_text += f"🚨 急募  \n"
            job_text += f"🔗 [詳細]({job['link']})  \n\n"
            if len(main_content) + len(job_text) > available_chars:
                skipped_count = len(jobs) - displayed_count
                print(f"📝 文字数制限により {displayed_count}件表示、{skipped_count}件スキップ")
                break
            main_content += job_text
            displayed_count += 1
        final_text = header_text + main_content + (f"\n📋 残り{skipped_count}件の案件はJSONファイルで確認できます。" if skipped_count > 0 else footer_text)
        actual_chars = len(final_text)
        print(f"📊 実際の文字数: {actual_chars:,}文字 / {MAX_CHARS:,}文字 ({actual_chars/MAX_CHARS*100:.1f}%)")
        print(f"📊 表示案件数: {displayed_count}件 / {len(jobs)}件")
        return {
            "@type": "MessageCard",
            "@context": "https://schema.org/extensions",
            "summary": f"Lancers全案件 {len(jobs)}件",
            "themeColor": "0078D4",
            "title": f"🚀 Lancers全案件リスト ({len(jobs)}件発見 / {displayed_count}件表示) - {datetime.now().strftime('%Y/%m/%d %H:%M')}",
            "text": final_text,
            "potentialAction": [{
                "@type": "OpenUri",
                "name": "🔍 Lancersで案件を探す",
                "targets": [{"os": "default", "uri": LANCERS_SEARCH_URL}]
            }]
        }

    async def send_to_teams(self, jobs):
        try:
            from dotenv import load_dotenv
            load_dotenv()
        except:
            pass
        webhook_url = os.getenv("TEAMS_WEBHOOK_URL")
        if not webhook_url:
            print("❌ Teams Webhook URLが設定されていません")
            return False
        payload = self.create_teams_payload(jobs)
        try:
            async with aiohttp.ClientSession() as session:
                print("📤 Teamsに全案件リストを送信中...")
                async with session.post(
                    webhook_url,
                    json=payload,
                    headers={"Content-Type": "application/json"}
                ) as response:
                    if response.status == 200:
                        print("✅ Teams送信成功！")
                        return True
                    else:
                        print(f"❌ Teams送信失敗 (ステータス: {response.status})")
                        return False
        except Exception as e:
            print(f"❌ Teams送信エラー: {e}")
            return False

    def save_data(self, jobs):
        timestamp = datetime.now()
        data = {
            "timestamp": timestamp.isoformat(),
            "count": len(jobs),
            "type": "全案件リスト",
            "skill_summary": self.create_skill_summary(jobs),
            "skill_distribution": self.create_skill_distribution(jobs),
            "jobs": jobs
        }
        filename = f"all_jobs_{timestamp.strftime('%Y%m%d_%H%M')}.json"
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"💾 全案件データを保存: {filename}")

    def create_skill_summary(self, jobs):
        skill_counts = {}
        for job in jobs:
            for match in job["skill_matches"]:
                skill = match["skill"]
                skill_counts[skill] = skill_counts.get(skill, 0) + 1
        return dict(sorted(skill_counts.items(), key=lambda x: x[1], reverse=True))

    def create_skill_distribution(self, jobs):
        total_jobs = len(jobs)
        no_skill_jobs = len([job for job in jobs if job["skill_count"] == 0])
        multi_skill_jobs = len([job for job in jobs if job["skill_count"] >= 2])
        high_priority_jobs = len([job for job in jobs if job["priority_score"] >= 10])
        return {
            "total_jobs": total_jobs,
            "no_skill_match": no_skill_jobs,
            "multi_skill_match": multi_skill_jobs,
            "high_priority": high_priority_jobs,
            "skill_match_rate": round((total_jobs - no_skill_jobs) / total_jobs * 100, 1) if total_jobs > 0 else 0
        }
        

# =============================
# Excel クリーニング
# =============================
def clean_excel_data(excel_path: str = EXCEL_PATH):
    """Excelファイル内の重複データと期限切れデータをクリーニング"""
    try:
        path = Path(excel_path)
        if not path.exists():
            print("📄 Excelファイルが存在しません（初回など）")
            return

        wb = openpyxl.load_workbook(path)
        if "ランサーズ" not in wb.sheetnames:
            print("📄 ランサーズシートが存在しません")
            return

        ws = wb["ランサーズ"]

        # データ読み込み
        all_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, 9):  # 8列
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            if len(row_data) > 5 and row_data[5]:
                all_rows.append({
                    'row_index': row_idx,
                    'date': row_data[0],
                    'title': row_data[1],
                    'category': row_data[2],
                    'price': row_data[3],
                    'deadline': row_data[4],
                    'url': row_data[5],
                    'score': row_data[6] if len(row_data) > 6 else None,
                    'skills': row_data[7] if len(row_data) > 7 else None
                })

        print(f"📊 処理前のデータ数: {len(all_rows)}件")

        # 1) URL重複（新しいものを残す）
        url_latest = {}
        for row in all_rows:
            url = row['url']
            if url not in url_latest:
                url_latest[url] = row
            else:
                try:
                    current_date = datetime.strptime(str(row['date']), "%Y-%m-%d %H:%M:%S")
                    existing_date = datetime.strptime(str(url_latest[url]['date']), "%Y-%m-%d %H:%M:%S")
                    if current_date > existing_date:
                        url_latest[url] = row
                except:
                    pass

        # 2) 期限切れ/古いデータ
        now = datetime.now()
        one_month_ago = now - timedelta(days=30)

        filtered_rows = []
        removed_count = {'duplicate': len(all_rows) - len(url_latest), 'expired': 0, 'old': 0}

        for url, row in url_latest.items():
            keep_row = True
            try:
                row_date = datetime.strptime(str(row['date']), "%Y-%m-%d %H:%M:%S")
                if row_date < one_month_ago:
                    keep_row = False
                    removed_count['old'] += 1
            except:
                pass

            if keep_row and row['deadline']:
                deadline_text = str(row['deadline'])
                if '締切' in deadline_text or '終了' in deadline_text:
                    keep_row = False
                    removed_count['expired'] += 1
                else:
                    try:
                        deadline_date = None
                        date_match = re.search(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', deadline_text)
                        if not date_match:
                            date_match = re.search(r'(\d{1,2})[/-](\d{1,2})', deadline_text)
                            if date_match:
                                month, day = int(date_match.group(1)), int(date_match.group(2))
                                year = now.year
                                deadline_date = datetime(year, month, day)
                            else:
                                date_match = re.search(r'(\d{1,2})月(\d{1,2})日', deadline_text)
                                if date_match:
                                    month, day = int(date_match.group(1)), int(date_match.group(2))
                                    year = now.year
                                    deadline_date = datetime(year, month, day)
                        else:
                            year, month, day = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
                            deadline_date = datetime(year, month, day)

                        if deadline_date and deadline_date < now:
                            keep_row = False
                            removed_count['expired'] += 1
                    except:
                        pass

            if keep_row:
                filtered_rows.append(row)

        print(f"🗑️ 削除データ: 重複 {removed_count['duplicate']} / 1ヶ月以上前 {removed_count['old']} / 期限切れ {removed_count['expired']}")
        print(f"📊 処理後のデータ数: {len(filtered_rows)}件")

        # 再作成
        wb.remove(ws)
        ws = wb.create_sheet("ランサーズ", 0)
        ws.append(["取得日時","タイトル","カテゴリ","価格","締切","URL","優先度スコア","スキル概要"])

        for row in sorted(filtered_rows, key=lambda x: str(x['date']), reverse=True):
            ws.append([row['date'], row['title'], row['category'], row['price'], row['deadline'], row['url'], row['score'], row['skills']])
            last = ws.max_row
            if row['url']:
                try:
                    ws.cell(row=last, column=6).hyperlink = row['url']
                    ws.cell(row=last, column=6).style = "Hyperlink"
                except:
                    pass

        _autosize_columns(ws)
        wb.save(path)
        print(f"✅ クリーニング完了: {path}")

        return {'before': len(all_rows), 'after': len(filtered_rows), 'removed': removed_count}

    except Exception as e:
        print(f"❌ クリーニングエラー: {e}")
        import traceback
        traceback.print_exc()
        return None

# =============================
# エントリポイント
# =============================
async def main():
    print("=" * 70)
    print("🤖 Lancers全案件取得システム（Teams28KB最大活用版）")
    print("=" * 70)

    print("\n📧 既存データのクリーニング中...")
    clean_result = clean_excel_data(EXCEL_PATH)
    if clean_result:
        print(f"   処理前: {clean_result['before']}件 → 処理後: {clean_result['after']}件")

    # ...続く処理もすべてインデントしておくこと...

    notifier = CompleteJobsNotifier()
    jobs = await notifier.fetch_jobs()

    if jobs:
        # JSON保存（リポジトリ or ローカル）
        notifier.save_data(jobs)

        # 🔧 ここで先に定義する！
        excel_data = {
            "timestamp": datetime.now().isoformat(),
            "count": len(jobs),
            "type": "全案件リスト",
            "skill_summary": notifier.create_skill_summary(jobs),
            "skill_distribution": notifier.create_skill_distribution(jobs),
            "jobs": jobs
        }

        # Excelの「ランサーズ」シートを上書き
        replace_lancers_sheet(excel_data, EXCEL_PATH)

        # 追記後クリーニング
        print("\n📧 追記後のクリーニング...")
        clean_excel_data(EXCEL_PATH)

        # Teams送信
        teams_success = await notifier.send_to_teams(jobs)

        print("\n" + "=" * 70)
        print("📊 実行結果:")
        print("=" * 70)
        print(f"✅ 全案件数: {len(jobs)}件")
        print(f"📤 Teams送信: {'成功' if teams_success else '失敗'}")
        print(f"💾 データ保存: 完了")

        skill_summary = notifier.create_skill_summary(jobs)
        skill_distribution = notifier.create_skill_distribution(jobs)
        print(f"\n🔧 スキル別マッチ件数（上位10位）:")
        for skill, count in list(skill_summary.items())[:10]:
            print(f"   {skill}: {count}件")

        print(f"\n📈 スキル分布:")
        print(f"   スキルマッチ率: {skill_distribution['skill_match_rate']}%")
        print(f"   複数スキルマッチ: {skill_distribution['multi_skill_match']}件")
        print(f"   高優先度案件: {skill_distribution['high_priority']}件")
        print(f"   スキルマッチなし: {skill_distribution['no_skill_match']}件")
    else:
        print("❌ 案件が見つかりませんでした")


if __name__ == "__main__":
    asyncio.run(main())
